
import json
import os
import queue
import sys
import threading
from concurrent.futures import CancelledError, ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk

from tkinterdnd2 import DND_FILES, TkinterDnD

sys.path.insert(0, str(Path(__file__).parent))
from pdfparser import parse_scanned_vtu
from scraper import VTUScraperGUI, initialize_ocr


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


def _apply_window_icon(window: tk.Misc):
    try:
        ico_path = resource_path("combined.ico")
        png_path = resource_path("combined.png")

        if os.path.exists(ico_path):
            try:
                window.iconbitmap(ico_path)
            except tk.TclError:
                pass

        if os.path.exists(png_path):
            try:
                icon_img = tk.PhotoImage(file=png_path)
                window._combined_icon_ref = icon_img
                window.iconphoto(True, icon_img)
            except tk.TclError:
                pass
    except Exception:
        pass


def _get_config_path() -> Path:
    appdata = Path(os.environ.get("APPDATA") or Path.home())
    config_dir = appdata / "VTUResultsUtility"
    config_dir.mkdir(parents=True, exist_ok=True)
    return config_dir / "config.json"


W, H         = 700, 620
BG           = "#f5f7fa"
CARD         = "#ffffff"
ACCENT       = "#1a73e8"
ACCENT_DK    = "#1558b0"
ACCENT_PALE  = "#e8f0fe"
ACCENT_HOVER = "#d2e3fc"
T_DARK       = "#202124"
T_MID        = "#5f6368"
T_LIGHT      = "#9aa0a6"
BORDER_CLR   = "#dadce0"
OK_CLR       = "#1e8e3e"
ERR_CLR      = "#d93025"
WARN_CLR     = "#f9ab00"

F_TITLE  = ("Segoe UI", 17, "bold")
F_HEAD   = ("Segoe UI", 12, "bold")
F_BODY   = ("Segoe UI", 11)
F_SMALL  = ("Segoe UI", 10)
F_LOG    = ("Consolas", 10)

COLS = ["USN", "Name", "Subject Code", "Subject Name",
        "Internal", "External", "Total", "Result"]

INST_NAME      = "YENEPOYA INSTITUTE OF TECHNOLOGY"
DEPT_NAME      = "Department of CSE(Data Science)"
YEAR_PERIOD_DEFAULT = "2025-26"
REVAL_DEFAULT = "Before Revaluation"
SEM_DEFAULT = "I Sem"
MAX_SUBJ_MARKS = 100  
PASS_MARK      = 40   



def _thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _no_border():
    from openpyxl.styles import Border, Side
    s = Side(style=None)
    return Border(left=s, right=s, top=s, bottom=s)


def _result_sheet_title(cfg: dict[str, str]) -> str:
    return f"RESULT SHEET {cfg['year_period']} - {cfg['reval_status']} - {cfg['semester']}"


def _credit_sheet_title(cfg: dict[str, str]) -> str:
    return f"RESULT SHEET {cfg['year_period']} - {cfg['reval_status']} - Credit Points - {cfg['semester']}"


def _save_excel(
    rows: list,
    out_dir: Path,
    subject_credits: dict[str, int],
    cfg: dict[str, str],
    subject_order: list[str] | None = None,
    highlight_mapping: dict[str, bool] | None = None,
) -> Path:

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    def _is_fail_result(result_val: str) -> bool:
        res = str(result_val or "").strip().upper()
        if not res:
            return False
        return res != "P"

    def _name_row_height(name_value: str, col_width: float) -> float:
        text = str(name_value or "").strip()
        if not text:
            return 16
        line_capacity = max(10, int(col_width - 2))
        lines = 0
        for part in text.splitlines() or [text]:
            lines += max(1, (len(part) + line_capacity - 1) // line_capacity)
        return max(16, min(45, 15 * lines))

    sorted_rows = sorted(rows, key=lambda r: (r.get("USN", ""), r.get("Subject Code", "")))
    student_order = []
    student_names = {}
    student_subjs = {}      
    for r in sorted_rows:
        usn  = r.get("USN", "")
        code = r.get("Subject Code", "")
        if usn not in student_subjs:
            student_order.append(usn)
            student_names[usn] = str(r.get("Name", "")).upper()
            student_subjs[usn] = {}
        student_subjs[usn][code] = r

    detected_codes = sorted({r.get("Subject Code", "") for r in rows if r.get("Subject Code", "")})
    if subject_order:
        ordered_valid = [c for c in subject_order if c in detected_codes]
        missing = [c for c in detected_codes if c not in ordered_valid]
        all_codes = ordered_valid + missing
    else:
        all_codes = detected_codes
    n_subj    = len(all_codes)
    cl = get_column_letter
    COL_SNO = 1
    COL_USN = 2
    COL_NAME = 3
    white_fill  = PatternFill("solid", fgColor="FFFFFF")

    hdr_font    = Font(bold=True, color="000000", name="Segoe UI", size=9)
    sub_font    = Font(bold=True, color="000000", name="Segoe UI", size=8)
    data_font   = Font(name="Segoe UI", size=9)
    title_font  = Font(name="Segoe UI", size=11, bold=True, color="000000")

    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ctr_align  = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin       = _thin_border()
    no_border  = _no_border()

    wb = Workbook()

    def _write_title_block(ws, title_text: str, n_cols: int):
        faculty_text = cfg.get("faculty_incharge", "").strip()
        title_rows = [
            (cfg["inst_name"], 14, 24),
            (cfg["dept_name"], 11, 20),
            (title_text, 11, 20),
            (f"Faculty In-charge: {faculty_text}" if faculty_text else "", 10, 16),
        ]
        for r_i, (text, fsize, row_h) in enumerate(title_rows, 1):
            ws.merge_cells(start_row=r_i, start_column=1, end_row=r_i, end_column=n_cols)
            c = ws.cell(r_i, 1, text)
            c.fill = white_fill
            c.font = Font(name="Segoe UI", size=fsize, bold=True, color="000000")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = no_border
            ws.row_dimensions[r_i].height = row_h

    def _write_summary_block(
        ws,
        data_row: int,
        last_dr: int,
        summary_stats: list[tuple[int, int, int]],
        col_fail: int,
        total_students: int,
    ):
        summary_start = last_dr + 3
        summary_labels = [
            "SUBJECT",
            "NO. OF STUDENTS TAKING THE EXAM",
            "NO. OF STUDENTS PASS IN THE PAPER",
            "NO. OF STUDENTS FAIL IN THE PAPER",
            "RESULT IN PERCENTAGE (%)",
            "NO. OF STUDENTS PASSED IN ALL SUBJECTS",
            "NO OF STUDENTS FAIL",
            "RESULT IN PERCENTAGE (%)",
        ]

        summary_col0 = 4
        for idx, label in enumerate(summary_labels):
            row_no = summary_start + idx
            ws.merge_cells(start_row=row_no, start_column=1, end_row=row_no, end_column=3)
            c = ws.cell(row_no, 1, label)
            c.fill = white_fill
            c.font = hdr_font if idx == 0 else title_font
            c.alignment = left_align if idx > 0 else hdr_align
            c.border = thin
            for extra_col in range(2, 4):
                ws.cell(row_no, extra_col).fill = white_fill
                ws.cell(row_no, extra_col).border = thin

        for i, code in enumerate(all_codes):
            col = summary_col0 + i
            c = ws.cell(summary_start, col, code)
            c.fill = white_fill
            c.font = hdr_font
            c.alignment = ctr_align
            c.border = thin

            taking, passed, failed = summary_stats[i]
            pct = 0 if taking == 0 else (passed / taking) * 100
            ws.cell(summary_start + 1, col, taking)
            ws.cell(summary_start + 2, col, passed)
            ws.cell(summary_start + 3, col, failed)
            ws.cell(summary_start + 4, col, pct)

            for row_no in range(summary_start + 1, summary_start + 5):
                cell = ws.cell(row_no, col)
                cell.fill = white_fill
                cell.font = data_font
                cell.alignment = ctr_align
                cell.border = thin
            ws.cell(summary_start + 4, col).number_format = "0.00"

        overall_subject_span_end = summary_col0 + max(n_subj - 1, 0)
        if n_subj:
            ws.merge_cells(start_row=summary_start + 5, start_column=summary_col0,
                           end_row=summary_start + 5, end_column=overall_subject_span_end)
            ws.merge_cells(start_row=summary_start + 6, start_column=summary_col0,
                           end_row=summary_start + 6, end_column=overall_subject_span_end)
            ws.merge_cells(start_row=summary_start + 7, start_column=summary_col0,
                           end_row=summary_start + 7, end_column=overall_subject_span_end)

        ws.cell(summary_start + 5, summary_col0,
                f'=COUNTIF({cl(col_fail)}{data_row}:{cl(col_fail)}{last_dr},0)')
        ws.cell(summary_start + 6, summary_col0,
                f'=COUNTIF({cl(col_fail)}{data_row}:{cl(col_fail)}{last_dr},">0")')
        ws.cell(summary_start + 7, summary_col0,
                f"=IF({total_students}=0,0,{cl(summary_col0)}{summary_start + 5}/{total_students}*100)")

        for row_no in range(summary_start + 5, summary_start + 8):
            c = ws.cell(row_no, summary_col0)
            c.fill = white_fill
            c.font = data_font
            c.alignment = ctr_align
            c.border = thin
            if row_no == summary_start + 7:
                c.number_format = "0.00"
            for col in range(summary_col0 + 1, overall_subject_span_end + 1):
                ws.cell(row_no, col).fill = white_fill
                ws.cell(row_no, col).border = thin

    ws = wb.active
    ws.title = "Result Sheet"
    DATA_ROW = 7
    COL_S0 = 4
    col_total = COL_S0 + n_subj * 3
    col_pct = col_total + 1
    col_fail = col_pct + 1
    n_cols = col_fail

    _write_title_block(ws, _result_sheet_title(cfg), n_cols)

    for col, text in [(COL_SNO, "S.No"), (COL_USN, "USN"), (COL_NAME, "STUDENT NAME")]:
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        c = ws.cell(5, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for i, code in enumerate(all_codes):
        c0 = COL_S0 + i * 3
        ws.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 2)
        c = ws.cell(5, c0, code)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for col, text in [
        (col_total, "Total\nMarks"),
        (col_pct, "%"),
        (col_fail, "Total no.\nof Fail"),
    ]:
        ws.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        c = ws.cell(5, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin
    ws.row_dimensions[5].height = 30

    for i in range(n_subj):
        c0 = COL_S0 + i * 3
        for off, lbl in enumerate(("INT", "EXT", "TOT")):
            c = ws.cell(6, c0 + off, lbl)
            c.fill = white_fill; c.font = sub_font; c.alignment = ctr_align; c.border = thin
    ws.row_dimensions[6].height = 16

    tot_cols = [COL_S0 + i * 3 + 2 for i in range(n_subj)]

    for s_idx, usn in enumerate(student_order):
        dr = DATA_ROW + s_idx

        def wc(col, val, align=ctr_align, fmt=None, fill_color=white_fill, _dr=dr):
            c = ws.cell(_dr, col, val)
            c.fill = fill_color
            c.font = data_font
            c.alignment = align
            c.border = thin
            if fmt:
                c.number_format = fmt
            return c

        wc(COL_SNO, s_idx + 1)
        wc(COL_USN, usn)
        student_name = student_names.get(usn, "")
        wc(COL_NAME, student_name, left_align)

        for i, code in enumerate(all_codes):
            subj = student_subjs[usn].get(code, {})
            c0 = COL_S0 + i * 3
            
            is_highlighted_fail = False
            raw_ext = subj.get("External", "")
            if highlight_mapping and highlight_mapping.get(code, False):
                try:
                    ext_val = int(raw_ext)
                    if ext_val < 18:
                        is_highlighted_fail = True
                except (ValueError, TypeError):
                    pass
            
            cell_fill = PatternFill("solid", fgColor="FFCCCC") if is_highlighted_fail else white_fill
            
            for off, key in enumerate(("Internal", "External", "Total")):
                raw = subj.get(key, "")
                try:
                    val = int(raw) if raw != "" else ""
                except (ValueError, TypeError):
                    val = raw
                wc(c0 + off, val, fill_color=cell_fill)

        wc(col_total, "=" + "+".join(f"{cl(c)}{dr}" for c in tot_cols), fmt="0")
        wc(col_pct, f"={cl(col_total)}{dr}/{n_subj * MAX_SUBJ_MARKS}*100", fmt="0.00")
        fail_count = sum(
            1
            for code in all_codes
            if _is_fail_result(student_subjs[usn].get(code, {}).get("Result", ""))
        )
        wc(col_fail, fail_count)
        ws.row_dimensions[dr].height = _name_row_height(student_name, 30)

    last_dr = DATA_ROW + max(len(student_order) - 1, 0)
    result_summary_stats = []
    for code in all_codes:
        taking = 0
        passed = 0
        failed = 0
        for usn in student_order:
            res = str(student_subjs[usn].get(code, {}).get("Result", "")).strip().upper()
            if not res:
                continue
            taking += 1
            if _is_fail_result(res):
                failed += 1
            else:
                passed += 1
        result_summary_stats.append((taking, passed, failed))
    _write_summary_block(ws, DATA_ROW, last_dr, result_summary_stats, col_fail, len(student_order))

    ws.column_dimensions[cl(COL_SNO)].width = 5
    ws.column_dimensions[cl(COL_USN)].width = 14
    ws.column_dimensions[cl(COL_NAME)].width = 30
    for col in range(COL_S0, n_cols + 1):
        ws.column_dimensions[cl(col)].width = 10
    ws.freeze_panes = f"D{DATA_ROW}"

    ws_credit = wb.create_sheet("Credit Sheet")
    C_DATA_ROW = 7
    C_COL_S0 = 4
    c_col_total_cp = C_COL_S0 + n_subj * 3
    c_col_sgpa = c_col_total_cp + 1
    c_col_remarks = c_col_sgpa + 1
    c_col_backlogs = c_col_remarks + 1
    c_col_pct = c_col_backlogs + 1
    c_n_cols = c_col_pct
    total_credits = sum(subject_credits.get(code, 0) for code in all_codes)

    _write_title_block(ws_credit, _credit_sheet_title(cfg), c_n_cols)

    for col, text in [(COL_SNO, "S.No"), (COL_USN, "USN"), (COL_NAME, "STUDENT NAME")]:
        ws_credit.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        c = ws_credit.cell(5, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for i, code in enumerate(all_codes):
        c0 = C_COL_S0 + i * 3
        ws_credit.merge_cells(start_row=5, start_column=c0, end_row=5, end_column=c0 + 2)
        c = ws_credit.cell(5, c0, code)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin

    for col, text in [
        (c_col_total_cp, "TOTAL CP"),
        (c_col_sgpa, "SGPA"),
        (c_col_remarks, "REMARKS"),
        (c_col_backlogs, "No. of Backlogs"),
        (c_col_pct, "Percentage"),
    ]:
        ws_credit.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        c = ws_credit.cell(5, col, text)
        c.fill = white_fill; c.font = hdr_font; c.alignment = hdr_align; c.border = thin
    ws_credit.row_dimensions[5].height = 30

    for i in range(n_subj):
        c0 = C_COL_S0 + i * 3
        for off, lbl in enumerate(("TOT", "GP", "CP")):
            c = ws_credit.cell(6, c0 + off, lbl)
            c.fill = white_fill; c.font = sub_font; c.alignment = ctr_align; c.border = thin
    ws_credit.row_dimensions[6].height = 16

    credit_tot_cols = [C_COL_S0 + i * 3 for i in range(n_subj)]
    credit_gp_cols = [C_COL_S0 + i * 3 + 1 for i in range(n_subj)]
    credit_cp_cols = [C_COL_S0 + i * 3 + 2 for i in range(n_subj)]

    for s_idx, usn in enumerate(student_order):
        dr = C_DATA_ROW + s_idx

        def wc_credit(col, val, align=ctr_align, fmt=None, _dr=dr):
            c = ws_credit.cell(_dr, col, val)
            c.fill = white_fill
            c.font = data_font
            c.alignment = align
            c.border = thin
            if fmt:
                c.number_format = fmt
            return c

        wc_credit(COL_SNO, s_idx + 1)
        wc_credit(COL_USN, usn)
        student_name = student_names.get(usn, "")
        wc_credit(COL_NAME, student_name, left_align)

        for i, code in enumerate(all_codes):
            subj = student_subjs[usn].get(code, {})
            c0 = C_COL_S0 + i * 3
            raw_total = subj.get("Total", "")
            try:
                total_val = int(raw_total) if raw_total != "" else ""
            except (ValueError, TypeError):
                total_val = raw_total

            wc_credit(c0, total_val)

            tot_ref = f"{cl(c0)}{dr}"
            gp_ref = f"{cl(c0 + 1)}{dr}"
            credit = int(subject_credits.get(code, 0))
            gp_formula = (
                f'=IF(ISNUMBER({tot_ref}),IF({tot_ref}>=90,10,IF({tot_ref}>=80,9,'
                f'IF({tot_ref}>=70,8,IF({tot_ref}>=60,7,IF({tot_ref}>=55,6,'
                f'IF({tot_ref}>=50,5,IF({tot_ref}>=40,4,0))))))),"")'
            )
            wc_credit(c0 + 1, gp_formula)
            wc_credit(c0 + 2, f'=IF({gp_ref}="","",{gp_ref}*{credit})')

        wc_credit(c_col_total_cp,
                  "=" + "+".join(f"{cl(c)}{dr}" for c in credit_cp_cols),
                  fmt="0")
        wc_credit(c_col_sgpa,
                  f"=IF({total_credits}=0,0,{cl(c_col_total_cp)}{dr}/{total_credits})",
                  fmt="0.00")

        fail_count = sum(
            1
            for code in all_codes
            if _is_fail_result(student_subjs[usn].get(code, {}).get("Result", ""))
        )
        wc_credit(c_col_backlogs, fail_count)
        wc_credit(c_col_remarks,
                  f'=IF({cl(c_col_backlogs)}{dr}=0,"PASS","FAIL")')

        tot_sum_expr = "+".join(f"{cl(c)}{dr}" for c in credit_tot_cols)
        wc_credit(c_col_pct,
                  f"=({tot_sum_expr})/{n_subj * MAX_SUBJ_MARKS}*100",
                  fmt="0.00")
        ws_credit.row_dimensions[dr].height = _name_row_height(student_name, 30)

    c_last_dr = C_DATA_ROW + max(len(student_order) - 1, 0)
    _write_summary_block(ws_credit, C_DATA_ROW, c_last_dr, result_summary_stats, c_col_backlogs, len(student_order))

    ws_credit.column_dimensions[cl(COL_SNO)].width = 5
    ws_credit.column_dimensions[cl(COL_USN)].width = 14
    ws_credit.column_dimensions[cl(COL_NAME)].width = 30
    for col in range(C_COL_S0, c_n_cols + 1):
        ws_credit.column_dimensions[cl(col)].width = 10
    ws_credit.freeze_panes = f"D{C_DATA_ROW}"

    ws2 = wb.create_sheet("Raw Data")
    RAW_COLS = ["USN", "Name", "Subject Code", "Subject Name",
                "Internal", "External", "Total", "Result"]
    raw_widths = {
        "USN": 16, "Name": 26, "Subject Code": 14, "Subject Name": 42,
        "Internal": 10, "External": 10, "Total": 8, "Result": 8,
    }
    for ci, cn in enumerate(RAW_COLS, 1):
        c = ws2.cell(1, ci, cn)
        c.fill = white_fill
        c.font = Font(bold=True, color="000000", name="Segoe UI", size=10)
        c.alignment = hdr_align; c.border = thin
    ws2.row_dimensions[1].height = 28

    data_font2 = Font(name="Segoe UI", size=10)
    for ri, r in enumerate(sorted_rows, 1):
        rf = white_fill
        row_name = ""
        for ci, cn in enumerate(RAW_COLS, 1):
            val = r.get(cn, "")
            if cn == "Name":
                val = str(val).upper()
                row_name = val
            c = ws2.cell(ri + 1, ci, val)
            c.fill = rf; c.font = data_font2; c.border = thin
            c.alignment = left_align if cn in ("Name", "Subject Name") else ctr_align
        ws2.row_dimensions[ri + 1].height = _name_row_height(row_name, raw_widths["Name"])

    ws2.column_dimensions[cl(1)].width = raw_widths["USN"]
    ws2.column_dimensions[cl(2)].width = raw_widths["Name"]
    ws2.column_dimensions[cl(3)].width = raw_widths["Subject Code"]
    for ci in range(4, len(RAW_COLS) + 1):
        ws2.column_dimensions[cl(ci)].width = 10
    ws2.freeze_panes = "A2"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out   = out_dir / f"VTU_Results_{stamp}.xlsx"
    try:
        wb.save(out)
    except PermissionError:
        out = out_dir / f"VTU_Results_{datetime.now().strftime('%Y%m%d_%H%M%S')}_new.xlsx"
        wb.save(out)

    return out



class VTUParserApp(tk.Frame):

    def __init__(self, parent):
        super().__init__(parent, bg=BG)
        self.parent = parent
        self.root = parent.winfo_toplevel()
        _apply_window_icon(self.root)

        self._scraper_win: tk.Toplevel | None = None
        self._scraper_app: VTUScraperGUI | None = None
        self._auto_import_after_scrape = tk.BooleanVar(value=True)
        
        self.drop_target_register(DND_FILES)
        self.dnd_bind('<<Drop>>', self._on_drop)
        self.dnd_bind('<<DropEnter>>', self._drag_enter)
        self.dnd_bind('<<DropLeave>>', self._drag_leave)

        self._folder: Path | None = None
        self._last_output_excel: Path | None = None
        self._zone = self
        self._busy = False
        self._cancel_parse_requested = False
        self._parse_executor: ThreadPoolExecutor | None = None
        self._q: queue.Queue = queue.Queue()
        self._last_credits: dict[str, int] = {}
        self._last_sort_pos: dict[str, int] = {}

        self._cfg_vars: dict[str, tk.StringVar] = {
            "inst_name": tk.StringVar(value=INST_NAME),
            "dept_name": tk.StringVar(value=DEPT_NAME),
            "year_period": tk.StringVar(value=YEAR_PERIOD_DEFAULT),
            "reval_status": tk.StringVar(value=REVAL_DEFAULT),
            "semester": tk.StringVar(value=SEM_DEFAULT),
            "faculty_incharge": tk.StringVar(value=""),
        }
        self._load_config()

        self._build_ui()
        self._set_main_controls_enabled(True)
        self._poll_queue()
        threading.Thread(target=initialize_ocr, daemon=True).start()
        try:
            import pyi_splash
            pyi_splash.close()
        except ImportError:
            pass


    def _build_ui(self):
        self._hdr = tk.Frame(self, bg=ACCENT, height=58)
        self._hdr.pack(fill="x")
        self._hdr.pack_propagate(False)
        tk.Label(self._hdr, text="VTU Bulk PDF Parser",
                 font=F_TITLE, bg=ACCENT, fg="white").pack(side="left", padx=20)

        body_host = tk.Frame(self, bg=BG)
        body_host.pack(fill="both", expand=True, padx=18, pady=12)
        self._body_host = body_host

        self._body_canvas = tk.Canvas(body_host, bg=BG, highlightthickness=0, bd=0)
        self._body_scroll = ttk.Scrollbar(body_host, orient="vertical", command=self._body_canvas.yview)
        self._body_canvas.configure(yscrollcommand=self._body_scroll.set)
        self._body_scroll.pack(side="right", fill="y")
        self._body_canvas.pack(side="left", fill="both", expand=True)

        body = tk.Frame(self._body_canvas, bg=BG)
        self._body_inner = body
        self._body_canvas_window = self._body_canvas.create_window((0, 0), window=body, anchor="nw")

        def _sync_scroll_region(_event=None):
            self._body_canvas.configure(scrollregion=self._body_canvas.bbox("all"))
            self._refresh_parser_scrollbar()

        def _sync_body_width(event):
            self._body_canvas.itemconfigure(self._body_canvas_window, width=event.width)
            _sync_scroll_region()

        body.bind("<Configure>", _sync_scroll_region)
        self._body_canvas.bind("<Configure>", _sync_body_width)

        def _on_mousewheel(event):
            self._body_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self._body_canvas.bind("<Enter>", lambda _e: self._body_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self._body_canvas.bind("<Leave>", lambda _e: self._body_canvas.unbind_all("<MouseWheel>"))

        self.after(120, self._fit_parser_window)

        ctrl = tk.Frame(body, bg=BG)
        ctrl.pack(fill="x", pady=(12, 0))

        self._btn_browse = tk.Button(
            ctrl, text="  Browse Folder  ", font=F_HEAD,
            bg=ACCENT, fg="white", activebackground=ACCENT_DK,
            activeforeground="white", bd=0, padx=16, pady=8,
            relief="flat", cursor="hand2", command=self._browse
        )
        self._btn_browse.pack(side="left")

        info = tk.Frame(ctrl, bg=BG)
        info.pack(side="left", padx=14)
        self._lbl_path  = tk.Label(info, text="No folder selected.",
                                   font=F_BODY, bg=BG, fg=T_MID,
                                   anchor="w", wraplength=430, justify="left")
        self._lbl_path.pack(anchor="w")
        self._lbl_count = tk.Label(info, text="",
                                   font=F_SMALL, bg=BG, fg=T_MID, anchor="w")
        self._lbl_count.pack(anchor="w")

        scraper_card = tk.Frame(body, bg=CARD, highlightbackground=BORDER_CLR, highlightthickness=1)
        scraper_card.pack(fill="x", pady=(12, 8))

        self._btn_open_scraper = tk.Button(
            scraper_card,
            text="OPEN VTU SCRAPER",
            font=("Segoe UI", 15, "bold"),
            bg=ACCENT,
            fg="white",
            activebackground=ACCENT_DK,
            activeforeground="white",
            bd=0,
            padx=22,
            pady=12,
            relief="flat",
            cursor="hand2",
            command=self._open_scraper,
        )
        self._btn_open_scraper.pack(fill="x", padx=14, pady=(12, 8))

        self._chk_auto_import = tk.Checkbutton(
            scraper_card,
            text="Import folder and auto start parsing",
            variable=self._auto_import_after_scrape,
            font=F_BODY,
            bg=CARD,
            fg=T_DARK,
            activebackground=CARD,
            activeforeground=T_DARK,
            selectcolor=CARD,
            anchor="w",
        )
        self._chk_auto_import.pack(fill="x", padx=14, pady=(0, 12))

        ttk.Separator(body, orient="horizontal").pack(fill="x", pady=12)

        cfg_card = tk.Frame(body, bg=CARD, highlightbackground=BORDER_CLR, highlightthickness=1)
        cfg_card.pack(fill="x", pady=(0, 10))

        tk.Label(
            cfg_card,
            text="Sheet Configuration",
            font=F_HEAD,
            bg=CARD,
            fg=T_DARK,
            anchor="w",
        ).grid(row=0, column=0, columnspan=6, sticky="w", padx=10, pady=(8, 6))

        tk.Label(cfg_card, text="College", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=1, column=0, sticky="w", padx=(14, 6), pady=(8, 4))
        ttk.Entry(cfg_card, textvariable=self._cfg_vars["inst_name"], font=F_BODY, width=32).grid(row=1, column=1, sticky="w", padx=(0, 14), pady=(8, 4))

        tk.Label(cfg_card, text="Department", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=1, column=2, sticky="w", padx=(6, 6), pady=(8, 4))
        ttk.Entry(cfg_card, textvariable=self._cfg_vars["dept_name"], font=F_BODY, width=30).grid(row=1, column=3, sticky="w", padx=(0, 14), pady=(8, 4))

        tk.Label(cfg_card, text="Year Period", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=2, column=0, sticky="w", padx=(14, 6), pady=(4, 8))
        ttk.Entry(cfg_card, textvariable=self._cfg_vars["year_period"], font=F_BODY, width=14).grid(row=2, column=1, sticky="w", padx=(0, 14), pady=(4, 8))

        tk.Label(cfg_card, text="Revaluation", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=2, column=2, sticky="w", padx=(6, 6), pady=(4, 8))
        reval_box = ttk.Combobox(
            cfg_card,
            textvariable=self._cfg_vars["reval_status"],
            state="readonly",
            font=F_BODY,
            width=18,
            values=["Before Revaluation", "After Revaluation"],
        )
        reval_box.grid(row=2, column=3, sticky="w", padx=(0, 14), pady=(4, 8))
       
        tk.Label(cfg_card, text="Semester", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=3, column=0, sticky="w", padx=(14, 6), pady=(4, 4))
        ttk.Entry(cfg_card, textvariable=self._cfg_vars["semester"], font=F_BODY, width=10).grid(row=3, column=1, sticky="w", padx=(0, 14), pady=(4, 4))

        tk.Label(cfg_card, text="Faculty In-charge", font=F_SMALL, bg=CARD, fg=T_MID).grid(row=3, column=2, sticky="w", padx=(6, 6), pady=(4, 4))
        ttk.Entry(cfg_card, textvariable=self._cfg_vars["faculty_incharge"], font=F_BODY, width=28).grid(row=3, column=3, sticky="w", padx=(0, 14), pady=(4, 4))

        tk.Button(
            cfg_card, text="  Save  ", font=F_SMALL,
            bg=ACCENT, fg="white", activebackground=ACCENT_DK, activeforeground="white",
            bd=0, padx=12, pady=4, relief="flat", cursor="hand2",
            command=self._save_config,
        ).grid(row=4, column=0, columnspan=4, sticky="e", padx=14, pady=(4, 10))

        self._progress_var = tk.DoubleVar(value=0)
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TCombobox", font=F_BODY)
        style.configure("Blue.Horizontal.TProgressbar",
                        troughcolor=BORDER_CLR, bordercolor=BORDER_CLR,
                        background=ACCENT, lightcolor=ACCENT, darkcolor=ACCENT_DK)
        pb_frame = tk.Frame(body, bg=BG)
        pb_frame.pack(fill="x")
        ttk.Progressbar(pb_frame, length=W - 52, mode="determinate",
                        maximum=100, variable=self._progress_var,
                        style="Blue.Horizontal.TProgressbar").pack(fill="x")

        log_outer = tk.Frame(body, bg=CARD,
                             highlightbackground=BORDER_CLR, highlightthickness=1)
        log_outer.pack(fill="both", expand=True, pady=(8, 0))

        self._log = tk.Text(log_outer, height=10, font=F_LOG, bg=CARD,
                            fg=T_DARK, bd=0, wrap="word",
                            state="disabled", cursor="arrow",
                            selectbackground=ACCENT_PALE)
        sb = ttk.Scrollbar(log_outer, command=self._log.yview)
        self._log.configure(yscrollcommand=sb.set)
        self._log.pack(side="left", fill="both", expand=True, padx=8, pady=6)
        sb.pack(side="right", fill="y")

        self._log.tag_configure("ok",   foreground=OK_CLR)
        self._log.tag_configure("err",  foreground=ERR_CLR)
        self._log.tag_configure("warn", foreground=WARN_CLR)
        self._log.tag_configure("info", foreground=T_MID)
        self._log.tag_configure("done", foreground=ACCENT,
                                font=("Segoe UI", 10, "bold"))

        btn_row = tk.Frame(body, bg=BG)
        btn_row.pack(pady=(10, 0))

        self._btn_parse = tk.Button(
            btn_row, text="  Parse & Export to Excel  ",
            font=("Segoe UI", 11, "bold"),
            bg=ACCENT, fg="white", activebackground=ACCENT_DK,
            activeforeground="white", bd=0, padx=28, pady=12,
            relief="flat", cursor="hand2", state="disabled",
            command=self._start_parse
        )
        self._btn_parse.pack(side="left", padx=(0, 8))

        self._btn_cancel_parse = tk.Button(
            btn_row,
            text="  Cancel Parsing  ",
            font=("Segoe UI", 11, "bold"),
            bg=T_LIGHT,
            fg="white",
            activebackground=ERR_CLR,
            activeforeground="white",
            bd=0,
            padx=20,
            pady=12,
            relief="flat",
            cursor="hand2",
            state="disabled",
            command=self._request_cancel_parse,
        )
        self._btn_cancel_parse.pack(side="left")

        util_row = tk.Frame(body, bg=BG)
        util_row.pack(fill="x", pady=(10, 0))

        self._btn_open_output_folder = tk.Button(
            util_row,
            text="Open Output Folder",
            font=F_BODY,
            bg=ACCENT_PALE,
            fg=T_DARK,
            activebackground=ACCENT_HOVER,
            activeforeground=T_DARK,
            bd=0,
            padx=16,
            pady=8,
            relief="flat",
            cursor="hand2",
            state="disabled",
            command=self._open_output_folder,
        )
        self._btn_open_output_folder.pack(side="left")

        self._btn_open_output_excel = tk.Button(
            util_row,
            text="Open Output Excel",
            font=F_BODY,
            bg=ACCENT_PALE,
            fg=T_DARK,
            activebackground=ACCENT_HOVER,
            activeforeground=T_DARK,
            bd=0,
            padx=16,
            pady=8,
            relief="flat",
            cursor="hand2",
            state="disabled",
            command=self._open_output_excel,
        )
        self._btn_open_output_excel.pack(side="left", padx=(10, 0))

    def _refresh_parser_scrollbar(self):
        try:
            self.update_idletasks()
            content_h = self._body_inner.winfo_reqheight()
            canvas_h = self._body_canvas.winfo_height()
            if content_h <= max(1, canvas_h):
                if self._body_scroll.winfo_ismapped():
                    self._body_scroll.pack_forget()
            else:
                if not self._body_scroll.winfo_ismapped():
                    self._body_scroll.pack(side="right", fill="y")
        except tk.TclError:
            return

    def _fit_parser_window(self):
        try:
            self.update_idletasks()
            screen_h = self.root.winfo_screenheight()
            screen_w = self.root.winfo_screenwidth()

            target_w = max(760, min(980, self._body_inner.winfo_reqwidth() + 70))
            target_h = self._hdr.winfo_reqheight() + self._body_inner.winfo_reqheight() + 48

            max_h = int(screen_h * 0.90)
            min_h = 650
            final_h = max(min_h, min(max_h, target_h))
            final_w = max(720, min(int(screen_w * 0.92), target_w))

            if self.root.winfo_exists():
                cur_geo = self.root.winfo_geometry()
                if "x" in cur_geo:
                    x = max(20, (screen_w - final_w) // 2)
                    y = max(20, (screen_h - final_h) // 2)
                    self.root.geometry(f"{final_w}x{final_h}+{x}+{y}")

            self.after(50, self._refresh_parser_scrollbar)
        except tk.TclError:
            return


    def _set_zone_bg(self, color):
        self._zone.configure(bg=color)
        for w in self._zone.winfo_children():
            w.configure(bg=color)

    def _drag_enter(self, event):
        self._set_zone_bg(ACCENT_HOVER)

    def _drag_leave(self, event):
        self._set_zone_bg(ACCENT_PALE)

    def _on_drop(self, event):
        self._set_zone_bg(ACCENT_PALE)
        try:
            paths = self.tk.splitlist(event.data.strip())
        except Exception:
            paths = [event.data.strip().strip("{}")]

        for raw in paths:
            path = Path(raw)
            if path.is_dir():
                self._set_folder(path)
                return
        self._log_msg("Dropped item is not a folder — please drop a folder.", "err")


    def _browse(self):
        if self._busy:
            return
        folder = filedialog.askdirectory(
            title="Select folder containing VTU PDF marksheets"
        )
        if folder:
            self._set_folder(Path(folder))

    def _set_main_controls_enabled(self, enabled: bool):
        state = "normal" if enabled else "disabled"
        if hasattr(self, "_btn_browse"):
            self._btn_browse.configure(state=state)
        if hasattr(self, "_btn_open_scraper"):
            self._btn_open_scraper.configure(state=state)
        if hasattr(self, "_chk_auto_import"):
            self._chk_auto_import.configure(state=state)

        if hasattr(self, "_btn_open_output_folder"):
            self._btn_open_output_folder.configure(state=state if self._folder else "disabled")

        if hasattr(self, "_btn_open_output_excel"):
            has_excel = self._last_output_excel is not None and self._last_output_excel.exists()
            self._btn_open_output_excel.configure(state=state if has_excel else "disabled")

    def _refresh_parse_controls(self):
        if not hasattr(self, "_btn_parse") or not hasattr(self, "_btn_cancel_parse"):
            return

        scraper_window_open = self._scraper_win is not None and self._scraper_win.winfo_exists()
        has_pdfs = bool(self._folder and self._folder.exists() and any(self._folder.glob("*.pdf")))

        if self._busy:
            self._btn_parse.configure(state="disabled", text="  Processing...  ")
            self._btn_cancel_parse.configure(state="normal", bg=ERR_CLR)
            return

        self._btn_cancel_parse.configure(state="disabled", bg=T_LIGHT)
        if scraper_window_open:
            self._btn_parse.configure(state="disabled", text="  Parse & Export to Excel  ")
        else:
            self._btn_parse.configure(
                state="normal" if has_pdfs else "disabled",
                text="  Parse & Export to Excel  ",
            )

    def _request_cancel_parse(self):
        if not self._busy:
            return
        self._cancel_parse_requested = True
        self._btn_cancel_parse.configure(state="disabled", bg=T_LIGHT)
        self._log_msg("Cancellation requested. Stopping parser workers...", "warn")

    def _open_output_folder(self):
        if not self._folder or not self._folder.exists() or not self._folder.is_dir():
            messagebox.showwarning("Folder Missing", "Output folder is not available.")
            return
        try:
            os.startfile(str(self._folder))
        except Exception as exc:
            messagebox.showerror("Open Folder Failed", f"Could not open folder:\n{exc}")

    def _open_output_excel(self):
        if not self._last_output_excel or not self._last_output_excel.exists():
            messagebox.showwarning("Excel Missing", "No exported Excel file is available yet.")
            return
        try:
            os.startfile(str(self._last_output_excel))
        except Exception as exc:
            messagebox.showerror("Open Excel Failed", f"Could not open Excel file:\n{exc}")

    def _open_scraper(self):
        if self._busy:
            messagebox.showwarning(
                "Parser Busy",
                "Parsing is currently in progress. Please wait for it to finish before opening the scraper.",
            )
            return

        if self._scraper_win is not None and self._scraper_win.winfo_exists():
            self._scraper_win.lift()
            self._scraper_win.focus_force()
            return

        win = tk.Toplevel(self.root)
        win.title("VTU Auto-Scraper")
        win.configure(bg=BG)
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        win_w = max(600, min(700, int(screen_w * 0.36)))
        win_h = max(540, min(640, int(screen_h * 0.68)))
        pos_x = max(40, (screen_w - win_w) // 2)
        pos_y = max(30, (screen_h - win_h) // 2)
        win.geometry(f"{win_w}x{win_h}+{pos_x}+{pos_y}")
        win.minsize(600, 520)
        win.transient(self.root)
        win.grab_set()
        win.focus_force()
        _apply_window_icon(win)

        self._scraper_win = win
        self._scraper_app = VTUScraperGUI(
            win,
            on_complete_callback=self._on_scraper_complete,
            on_cancel_callback=self._on_scraper_cancel,
        )
        self._scraper_app.pack(fill="both", expand=True)

        # Prevent launching duplicate modal windows while one is open.
        self._btn_open_scraper.configure(state="disabled")
        self._refresh_parse_controls()

        def _on_close_scraper():
            self._close_scraper_window()

        win.protocol("WM_DELETE_WINDOW", _on_close_scraper)

    def _close_scraper_window(self, force: bool = False):
        if (
            not force
            and
            self._scraper_app is not None
            and hasattr(self._scraper_app, "is_scraping")
            and self._scraper_app.is_scraping
        ):
            messagebox.showwarning(
                "Scraper Running",
                "Scraper is still running. Use CANCEL inside scraper and wait for shutdown before closing this window.",
            )
            return

        if self._scraper_win is not None and self._scraper_win.winfo_exists():
            try:
                self._scraper_win.grab_release()
            except tk.TclError:
                pass
            self._scraper_win.destroy()
        self._scraper_win = None
        self._scraper_app = None
        if not self._busy:
            self._btn_open_scraper.configure(state="normal")
        self._refresh_parse_controls()

    def _on_scraper_cancel(self):
        if self._busy:
            self._request_cancel_parse()
            self._log_msg("Scraper cancellation propagated to parser cancellation.", "warn")

    def _on_scraper_complete(self, output_dir: str):
        self._close_scraper_window(force=True)

        self._log_msg("Scraper finished. Using downloaded PDFs for parser.", "info")

        if not self._auto_import_after_scrape.get():
            try:
                os.startfile(output_dir)
                self._log_msg(f"Auto-import disabled. Opened scraper output folder: {output_dir}", "info")
            except Exception as exc:
                self._log_msg(f"Could not open scraper output folder: {exc}", "err")
            return

        folder = Path(output_dir)
        if not folder.exists():
            messagebox.showerror("Import Failed", f"Downloaded folder not found:\n{output_dir}")
            return

        self._set_folder(folder)
        self._log_msg("Auto-import enabled. Folder selected from scraper output.", "info")

        if self._busy:
            self._log_msg("Parser is already busy. Auto parse skipped.", "warn")
            return

        self._log_msg("Auto-starting parse now...", "info")
        self.after(120, self._start_parse)

    def _set_folder(self, path: Path):
        if not path.exists() or not path.is_dir():
            self._lbl_path.configure(text="Invalid folder selected.", fg=ERR_CLR)
            self._lbl_count.configure(text="", fg=T_MID)
            self._btn_parse.configure(state="disabled")
            self._btn_open_output_folder.configure(state="disabled")
            self._refresh_parse_controls()
            self._log_msg(f"Invalid folder path: {path}", "err")
            return

        self._folder = path
        self._btn_open_output_folder.configure(state="normal")
        pdfs = sorted(path.glob("*.pdf"))
        self._lbl_path.configure(text=str(path), fg=T_DARK)
        if pdfs:
            self._lbl_count.configure(
                text=f"{len(pdfs)} PDF file(s) ready to parse.", fg=OK_CLR
            )
            self._log_msg(f"Folder: {path}", "info")
            names = ", ".join(p.name for p in pdfs[:6])
            suffix = f" … (+{len(pdfs) - 6} more)" if len(pdfs) > 6 else ""
            self._log_msg(f"PDFs found: {names}{suffix}", "info")
        else:
            self._lbl_count.configure(text="No PDF files found in this folder.", fg=ERR_CLR)
            self._log_msg("No PDF files found in the selected folder.", "err")
        self._refresh_parse_controls()

    def _read_export_config(self) -> dict[str, str] | None:
        cfg = {
            "inst_name": self._cfg_vars["inst_name"].get().strip(),
            "dept_name": self._cfg_vars["dept_name"].get().strip(),
            "year_period": self._cfg_vars["year_period"].get().strip(),
            "reval_status": self._cfg_vars["reval_status"].get().strip(),
            "semester": self._cfg_vars["semester"].get().strip(),
            "faculty_incharge": self._cfg_vars["faculty_incharge"].get().strip(),
        }
        for key, label in [
            ("inst_name", "College"),
            ("dept_name", "Department"),
            ("year_period", "Year Period"),
            ("reval_status", "Revaluation"),
            ("semester", "Semester"),
        ]:
            if not cfg[key]:
                messagebox.showerror("Missing Configuration", f"{label} cannot be empty.")
                return None
        return cfg


    def _start_parse(self):
        if self._busy or not self._folder:
            return

        if self._scraper_app is not None and getattr(self._scraper_app, "is_scraping", False):
            messagebox.showwarning(
                "Scraper Running",
                "Scraper is still running. Wait until it completes or cancel it before parsing.",
            )
            return

        cfg = self._read_export_config()
        if not cfg:
            return
        pdfs = sorted(self._folder.glob("*.pdf"))
        self._cancel_parse_requested = False
        if not pdfs:
            messagebox.showwarning("No PDFs", "No PDF files found in the selected folder.")
            self._refresh_parse_controls()
            return
        self._refresh_parse_controls()
        self._busy = True
        self._set_main_controls_enabled(False)
        self._btn_parse.configure(state="disabled", text="  Processing…  ")
        self._progress_var.set(0)
        self._log_msg(f"\nStarting bulk parse of {len(pdfs)} file(s) …", "info")
        threading.Thread(target=self._worker,
                         args=(pdfs, cfg), daemon=True).start()

    def _worker(self, pdfs: list, cfg: dict[str, str]):
        all_rows = []
        errors   = []
        n        = len(pdfs)

        max_workers = max(1, min(3, n, os.cpu_count() or 1))
        self._q.put(("log", f"Parallel parsing enabled: {max_workers} worker(s)", "info"))

        def _parse_single(pdf_path: Path):
            try:
                rows = parse_scanned_vtu(str(pdf_path))
                return pdf_path.name, rows, None
            except Exception as exc:
                return pdf_path.name, [], str(exc)

        completed = 0
        executor = ThreadPoolExecutor(max_workers=max_workers)
        self._parse_executor = executor
        futures = {executor.submit(_parse_single, pdf): pdf for pdf in pdfs}

        try:
            for future in as_completed(futures):
                if self._cancel_parse_requested:
                    for f in futures:
                        f.cancel()
                    executor.shutdown(wait=False, cancel_futures=True)
                    self._q.put(("done_cancel", None))
                    return

                completed += 1
                try:
                    pdf_name, rows, err = future.result()
                except CancelledError:
                    continue

                self._q.put(("log", f"[{completed}/{n}]  {pdf_name}", "info"))
                self._q.put(("progress", int(completed / n * 100)))

                if err:
                    errors.append(pdf_name)
                    self._q.put(("log", f"        Error: {err}", "err"))
                elif rows:
                    all_rows.extend(rows)
                    self._q.put(("log", f"        {len(rows)} subject(s) extracted.", "ok"))
                else:
                    self._q.put(("log", "        No subjects found — check scan quality.", "warn"))
        finally:
            self._parse_executor = None
            if not self._cancel_parse_requested:
                executor.shutdown(wait=True)

        self._q.put(("progress", 100))

        if self._cancel_parse_requested:
            self._q.put(("done_cancel", None))
            return

        if all_rows:
            subjects = sorted({r.get("Subject Code", "") for r in all_rows if r.get("Subject Code", "")})
            self._q.put(("need_credits", {
                "rows": all_rows,
                "errors": errors,
                "total_files": n,
                "subjects": subjects,
                "cfg": cfg,
            }))
        else:
            self._q.put(("log",      "\nNo data extracted from any PDF.", "err"))
            self._q.put(("done_err", None))

    def _export_worker(self, payload: dict, credits: dict[str, int], subject_order: list[str], highlight_mapping: dict[str, bool]):
        try:
            if self._cancel_parse_requested:
                self._q.put(("done_cancel", None))
                return
            out = _save_excel(payload["rows"], self._folder, credits, payload["cfg"], subject_order, highlight_mapping)
            n = payload["total_files"]
            errors = payload["errors"]
            msg = (f"\nDone!  "
                   f"{len(payload['rows'])} row(s) from "
                   f"{n - len(errors)}/{n} file(s)\n"
                   f"   Saved: {out.name}")
            self._q.put(("log", msg, "done"))
            self._q.put(("done_ok", str(out)))
        except Exception as exc:
            self._q.put(("log", f"\nCould not write Excel: {exc}", "err"))
            self._q.put(("done_err", None))

    def _prompt_subject_credits(self, subjects: list[str], auto_disable_highlight: set | None = None) -> tuple[dict[str, int], list[str], dict[str, str], dict[str, bool]] | None:
        dlg = tk.Toplevel(self)
        dlg.title("Subject Credits")
        dlg.configure(bg=BG)
        dlg.transient(self)
        dlg.grab_set()
        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()
        dlg_w = max(740, min(900, int(screen_w * 0.58)))
        dlg_h = max(620, min(760, int(screen_h * 0.78)))
        dlg_x = max(20, (screen_w - dlg_w) // 2)
        dlg_y = max(20, (screen_h - dlg_h) // 2)
        dlg.geometry(f"{dlg_w}x{dlg_h}+{dlg_x}+{dlg_y}")
        dlg.minsize(740, 620)
        dlg.resizable(True, True)
        _apply_window_icon(dlg)

        frm = tk.Frame(dlg, bg=BG, padx=16, pady=12)
        frm.pack(fill="both", expand=True)

        tk.Label(
            frm,
            text="Configure subject credits and order (required before export):",
            font=F_HEAD,
            bg=BG,
            fg=T_DARK,
            anchor="w",
            justify="left",
        ).grid(row=0, column=0, columnspan=4, sticky="w", pady=(0, 6))

        tk.Label(
            frm,
            text=f"Subjects detected: {len(subjects)}",
            font=F_SMALL,
            bg=BG,
            fg=T_MID,
            anchor="w",
        ).grid(row=1, column=0, columnspan=4, sticky="w", pady=(0, 2))

        tk.Label(
            frm,
            text="How to use: pick exactly 2 rows in 'Combine?' and click 'Combine Selected'. Use Up/Down to reorder rows.",
            font=F_SMALL,
            bg=BG,
            fg=T_MID,
            anchor="w",
            justify="left",
            wraplength=620,
        ).grid(row=2, column=0, columnspan=4, sticky="w", pady=(0, 10))

        select_all_var = tk.BooleanVar(value=True)
        def toggle_select_all():
            val = select_all_var.get()
            for var in highlight_vars.values():
                var.set(val)

        table_host = tk.Frame(frm, bg=BG)
        table_host.grid(row=3, column=0, columnspan=5, sticky="nsew", pady=(0, 4))
        frm.grid_rowconfigure(3, weight=1)
        frm.grid_columnconfigure(0, weight=1)

        table_canvas = tk.Canvas(table_host, bg=BG, highlightthickness=0, bd=0)
        table_scroll = ttk.Scrollbar(table_host, orient="vertical", command=table_canvas.yview)
        table_canvas.configure(yscrollcommand=table_scroll.set)
        table_scroll.pack(side="right", fill="y")
        table_canvas.pack(side="left", fill="both", expand=True)

        table_frm = tk.Frame(table_canvas, bg=BG)
        table_canvas_window = table_canvas.create_window((0, 0), window=table_frm, anchor="nw")

        def _update_table_scroll(_event=None):
            table_canvas.configure(scrollregion=table_canvas.bbox("all"))
            content_h = table_frm.winfo_reqheight()
            canvas_h = table_canvas.winfo_height()
            if content_h <= max(1, canvas_h):
                if table_scroll.winfo_ismapped():
                    table_scroll.pack_forget()
            else:
                if not table_scroll.winfo_ismapped():
                    table_scroll.pack(side="right", fill="y")

        def _update_table_width(event):
            table_canvas.itemconfigure(table_canvas_window, width=event.width)
            _update_table_scroll()

        table_frm.bind("<Configure>", _update_table_scroll)
        table_canvas.bind("<Configure>", _update_table_width)

        tk.Label(table_frm, text="Move", font=F_SMALL, bg=BG, fg=T_MID, width=8, anchor="w").grid(row=0, column=0, sticky="w", padx=(0, 10))
        tk.Label(table_frm, text="Combine?", font=F_SMALL, bg=BG, fg=T_MID, width=10, anchor="w").grid(row=0, column=1, sticky="w", padx=(0, 10))
        tk.Label(table_frm, text="Subject", font=F_SMALL, bg=BG, fg=T_MID, width=22, anchor="w").grid(row=0, column=2, sticky="w", padx=(0, 10))
        tk.Label(table_frm, text="Credit", font=F_SMALL, bg=BG, fg=T_MID, width=8, anchor="w").grid(row=0, column=3, sticky="w", padx=(0, 10))
        cb_all = tk.Checkbutton(
            table_frm,
            text="Highlight rows where external < 18",
            variable=select_all_var,
            command=toggle_select_all,
            font=F_SMALL,
            bg=BG,
            fg=T_MID,
        )
        cb_all.grid(row=0, column=4, sticky="w")
        
        ordered_subjects = list(subjects)
        try:
            ordered_subjects.sort(key=lambda c: self._last_sort_pos.get(c, 999))
        except Exception:
            pass
            
        credits_vars: dict[str, tk.StringVar] = {}
        selected_vars: dict[str, tk.BooleanVar] = {}
        highlight_vars: dict[str, tk.BooleanVar] = {}
        combined_mapping: dict[str, str] = {}
        row_widgets: dict[str, dict[str, tk.Widget]] = {}
        for code in ordered_subjects:
            credits_vars[code] = tk.StringVar(value=str(self._last_credits.get(code, "")))
            selected_vars[code] = tk.BooleanVar(value=False)
            highlight_vars[code] = tk.BooleanVar(value=code not in (auto_disable_highlight or set()))

        def _focus_credit(code: str, delta: int):
            idx = ordered_subjects.index(code)
            new_idx = max(0, min(len(ordered_subjects) - 1, idx + delta))
            next_code = ordered_subjects[new_idx]
            entry = row_widgets[next_code]["credit"]
            entry.focus_set()
            if isinstance(entry, ttk.Entry):
                entry.icursor("end")
                entry.selection_range(0, "end")

        def _focus_checkbox(code: str, key: str, delta: int):
            idx = ordered_subjects.index(code)
            new_idx = max(0, min(len(ordered_subjects) - 1, idx + delta))
            next_code = ordered_subjects[new_idx]
            row_widgets[next_code][key].focus_set()

        def _move_subject(code: str, direction: int):
            idx = ordered_subjects.index(code)
            new_idx = idx + direction
            if new_idx < 0 or new_idx >= len(ordered_subjects):
                return
            ordered_subjects[idx], ordered_subjects[new_idx] = ordered_subjects[new_idx], ordered_subjects[idx]
            _refresh_row_positions()

        def _build_row_widgets(code: str):
            move_frm = tk.Frame(table_frm, bg=BG)
            up_btn = tk.Button(move_frm, text="▲", font=F_SMALL, width=2, command=lambda c=code: _move_subject(c, -1))
            up_btn.pack(side="top")
            down_btn = tk.Button(move_frm, text="▼", font=F_SMALL, width=2, command=lambda c=code: _move_subject(c, 1))
            down_btn.pack(side="top", pady=(2, 0))

            combine_chk = tk.Checkbutton(table_frm, variable=selected_vars[code], bg=BG)
            subject_lbl = tk.Label(table_frm, text=code, font=F_BODY, bg=BG, fg=T_DARK, width=22, anchor="w")
            credit_ent = ttk.Entry(table_frm, textvariable=credits_vars[code], width=8, font=F_BODY)
            highlight_chk = tk.Checkbutton(table_frm, variable=highlight_vars[code], bg=BG)

            credit_ent.bind("<Tab>", lambda e, c=code: (_focus_credit(c, 1), "break")[1])
            credit_ent.bind("<Shift-Tab>", lambda e, c=code: (_focus_credit(c, -1), "break")[1])
            credit_ent.bind("<Down>", lambda e, c=code: (_focus_credit(c, 1), "break")[1])
            credit_ent.bind("<Up>", lambda e, c=code: (_focus_credit(c, -1), "break")[1])
            combine_chk.bind("<Down>", lambda e, c=code: (_focus_checkbox(c, "combine", 1), "break")[1])
            combine_chk.bind("<Up>", lambda e, c=code: (_focus_checkbox(c, "combine", -1), "break")[1])
            highlight_chk.bind("<Down>", lambda e, c=code: (_focus_checkbox(c, "highlight", 1), "break")[1])
            highlight_chk.bind("<Up>", lambda e, c=code: (_focus_checkbox(c, "highlight", -1), "break")[1])

            row_widgets[code] = {
                "move": move_frm,
                "up": up_btn,
                "down": down_btn,
                "combine": combine_chk,
                "subject": subject_lbl,
                "credit": credit_ent,
                "highlight": highlight_chk,
            }

        def _refresh_row_positions():
            for idx, code in enumerate(ordered_subjects, start=1):
                widgets = row_widgets[code]
                widgets["move"].grid(row=idx, column=0, sticky="w", padx=(0, 10), pady=2)
                widgets["combine"].grid(row=idx, column=1, sticky="w", padx=(0, 10), pady=2)
                widgets["subject"].grid(row=idx, column=2, sticky="w", padx=(0, 10), pady=2)
                widgets["credit"].grid(row=idx, column=3, sticky="w", padx=(0, 10), pady=2)
                widgets["highlight"].grid(row=idx, column=4, sticky="w", pady=2)

                widgets["up"].configure(state="disabled" if idx == 1 else "normal")
                widgets["down"].configure(state="disabled" if idx == len(ordered_subjects) else "normal")
            _update_table_scroll()

        def _combine_selected():
            selected = [code for code in ordered_subjects if selected_vars[code].get()]
            if len(selected) != 2:
                messagebox.showerror("Combine Error", "Please select exactly 2 subjects to combine.", parent=dlg)
                return
                
            code1, code2 = selected
            combined_code = f"{code1}/{code2}"
            
            idx1 = ordered_subjects.index(code1)
            ordered_subjects.remove(code1)
            ordered_subjects.remove(code2)
            ordered_subjects.insert(idx1, combined_code)
            
            combined_mapping[code1] = combined_code
            combined_mapping[code2] = combined_code
            
            credits_vars[combined_code] = tk.StringVar(value=credits_vars[code1].get())
            selected_vars[combined_code] = tk.BooleanVar(value=False)
            highlight_vars[combined_code] = tk.BooleanVar(value=False)

            for c in (code1, code2):
                for widget in row_widgets[c].values():
                    if widget.winfo_exists():
                        widget.destroy()
                del row_widgets[c]
            
            del credits_vars[code1]
            del credits_vars[code2]
            del selected_vars[code1]
            del selected_vars[code2]
            del highlight_vars[code1]
            del highlight_vars[code2]
            
            _build_row_widgets(combined_code)
            _refresh_row_positions()

        for code in ordered_subjects:
            _build_row_widgets(code)
        _refresh_row_positions()

        result: tuple[dict[str, int], list[str], dict[str, str], dict[str, bool]] | None = None

        def on_submit():
            nonlocal result
            final_credits: dict[str, int] = {}
            for code in ordered_subjects:
                raw = credits_vars[code].get().strip()
                if raw == "":
                    messagebox.showerror("Missing Credit", f"Credit missing for {code}", parent=dlg)
                    return
                if not raw.isdigit():
                    messagebox.showerror("Invalid Credit", f"Credit must be a non-negative integer for {code}", parent=dlg)
                    return
                final_credits[code] = int(raw)

            final_highlights = {code: var.get() for code, var in highlight_vars.items()}
            result = (final_credits, ordered_subjects, combined_mapping, final_highlights)
            dlg.destroy()

        def on_cancel():
            dlg.destroy()

        btns = tk.Frame(frm, bg=BG)
        btns.grid(row=4, column=0, columnspan=5, sticky="e", pady=(12, 0))
        tk.Button(btns, text="Combine Selected", font=F_BODY, command=_combine_selected).pack(side="left", padx=(0, 20))
        tk.Button(btns, text="Cancel", font=F_BODY, command=on_cancel, width=10).pack(side="right")
        tk.Button(btns, text="Export", font=F_BODY, command=on_submit, width=10).pack(side="right", padx=(0, 8))

        dlg.protocol("WM_DELETE_WINDOW", on_cancel)
        self.wait_window(dlg)
        return result


    def _poll_queue(self):
        try:
            while True:
                item = self._q.get_nowait()
                kind = item[0]
                if kind == "log":
                    self._log_msg(item[1], item[2])
                elif kind == "progress":
                    self._progress_var.set(item[1])
                elif kind == "need_credits":
                    if self._cancel_parse_requested:
                        self._q.put(("done_cancel", None))
                        continue
                    payload = item[1]
                    # Detect subjects where every student has external == 0 (exam not conducted)
                    _subj_ext: dict[str, list[int]] = {}
                    for _r in payload["rows"]:
                        _code = _r.get("Subject Code", "")
                        if not _code:
                            continue
                        try:
                            _subj_ext.setdefault(_code, []).append(int(_r.get("External", "")))
                        except (ValueError, TypeError):
                            pass
                    auto_disable = {
                        c for c, vals in _subj_ext.items() if vals and all(v == 0 for v in vals)
                    }
                    picked = self._prompt_subject_credits(payload["subjects"], auto_disable)
                    if not picked:
                        self._log_msg("\nExport cancelled: credits were not provided.", "err")
                        self._busy = False
                        self._cancel_parse_requested = False
                        self._set_main_controls_enabled(True)
                        self._refresh_parse_controls()
                    else:
                        credits, subject_order, combined_mapping, highlight_mapping = picked
                        self._last_credits = credits
                        self._last_sort_pos = {code: idx for idx, code in enumerate(subject_order, start=1)}
                        
                        for row in payload["rows"]:
                            code = row.get("Subject Code", "")
                            if code in combined_mapping:
                                row["Subject Code"] = combined_mapping[code]
                                
                        self._log_msg("Credits captured. Preparing workbook ...", "info")
                        threading.Thread(
                            target=self._export_worker,
                            args=(payload, credits, subject_order, highlight_mapping),
                            daemon=True,
                        ).start()
                elif kind == "done_ok":
                    self._last_output_excel = Path(item[1])
                    self._busy = False
                    self._cancel_parse_requested = False
                    self._set_main_controls_enabled(True)
                    self._refresh_parse_controls()
                    self._btn_open_output_excel.configure(state="normal")
                    messagebox.showinfo(
                        "Export Complete",
                        f"All results saved to:\n\n{item[1]}"
                    )
                elif kind == "done_err":
                    self._busy = False
                    self._cancel_parse_requested = False
                    self._set_main_controls_enabled(True)
                    self._refresh_parse_controls()
                elif kind == "done_cancel":
                    self._busy = False
                    self._cancel_parse_requested = False
                    self._set_main_controls_enabled(True)
                    self._progress_var.set(0)
                    self._refresh_parse_controls()
                    self._log_msg("Parsing cancelled by user.", "warn")
        except queue.Empty:
            pass
        self.after(80, self._poll_queue)

    def can_close(self) -> bool:
        if self._busy:
            messagebox.showwarning(
                "Parser Busy",
                "Parsing/export is in progress. Please wait for completion before closing the application.",
            )
            return False
        if self._scraper_app is not None and getattr(self._scraper_app, "is_scraping", False):
            messagebox.showwarning(
                "Scraper Running",
                "Scraper is running. Cancel it and wait for shutdown before closing the application.",
            )
            return False
        return True


    def _load_config(self):
        path = _get_config_path()
        if not path.exists():
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for key in ("inst_name", "dept_name", "year_period", "reval_status", "semester", "faculty_incharge"):
                if key in data and key in self._cfg_vars:
                    self._cfg_vars[key].set(str(data[key]))
        except Exception:
            pass

    def _save_config(self):
        path = _get_config_path()
        try:
            data = {key: var.get().strip() for key, var in self._cfg_vars.items()}
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2)
            self._log_msg("Configuration saved.", "ok")
        except Exception as exc:
            messagebox.showerror("Save Failed", f"Could not save configuration:\n{exc}")

    def _log_msg(self, text: str, tag: str = "info"):
        try:
            print(text)
        except Exception:
            pass
        self._log.configure(state="normal")
        self._log.insert("end", text + "\n", tag)
        self._log.see("end")
        self._log.configure(state="disabled")


if __name__ == "__main__":
    root = TkinterDnD.Tk()
    root.title("VTU Bulk PDF Parser")
    root.geometry(f"{W}x{H}")
    app = VTUParserApp(root)
    app.pack(fill="both", expand=True)
    root.mainloop()
